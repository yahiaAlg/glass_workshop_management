# apps/backup/services.py
import io
import zipfile
from datetime import datetime
from django.core.files.base import ContentFile
from import_export.formats import base_formats
import tablib

class BackupService:
    def __init__(self, backup_job):
        self.backup_job = backup_job
        self.resources = self._get_resources()
    
    def _get_resources(self):
        """Get import-export resources for selected models"""
        resources = {}
        
        # Define resources for each model based on the provided resource files
        if self.backup_job.include_company:
            try:
                from apps.company.resources import CompanyResource
                resources['company'] = CompanyResource()
            except ImportError:
                self.backup_job.add_log("Company resource not found")
        
        if self.backup_job.include_customers:
            try:
                from apps.customers.resources import CustomerResource
                resources['customers'] = CustomerResource()
            except ImportError:
                self.backup_job.add_log("Customer resource not found")
        
        if self.backup_job.include_products:
            try:
                from apps.inventory.resources import GlassProductResource
                resources['products'] = GlassProductResource()
            except ImportError:
                self.backup_job.add_log("Product resource not found")
        
        if self.backup_job.include_suppliers:
            try:
                from apps.suppliers.resources import SupplierResource
                resources['suppliers'] = SupplierResource()
            except ImportError:
                self.backup_job.add_log("Supplier resource not found")
        
        if self.backup_job.include_orders:
            try:
                from apps.orders.resources import OrderResource, OrderItemResource
                resources['orders'] = OrderResource()
                resources['order_items'] = OrderItemResource()
            except ImportError:
                self.backup_job.add_log("Order resources not found")
        
        if self.backup_job.include_invoices:
            try:
                from apps.invoices.resources import (
                    InvoiceResource, InvoiceItemResource, 
                    InvoiceServiceResource, PaymentResource
                )
                resources['invoices'] = InvoiceResource()
                resources['invoice_items'] = InvoiceItemResource()
                resources['invoice_services'] = InvoiceServiceResource()
                resources['payments'] = PaymentResource()
            except ImportError:
                self.backup_job.add_log("Invoice resources not found")
        
        # Add audit resources if available
        if hasattr(self.backup_job, 'include_audit') and self.backup_job.include_audit:
            try:
                from apps.audit.resources import (
                    ExpenseCategoryResource, RevenueCategoryResource,
                    AdditionalExpenseResource, AdditionalRevenueResource
                )
                resources['expense_categories'] = ExpenseCategoryResource()
                resources['revenue_categories'] = RevenueCategoryResource()
                resources['additional_expenses'] = AdditionalExpenseResource()
                resources['additional_revenues'] = AdditionalRevenueResource()
            except ImportError:
                self.backup_job.add_log("Audit resources not found")
        
        return resources
    
    def export_data(self):
        """Export data using django-import-export"""
        try:
            self.backup_job.status = 'processing'
            self.backup_job.save()
            
            format_map = {
                'json': base_formats.JSON(),
                'excel': base_formats.XLSX(),
                'csv': base_formats.CSV()
            }
            
            format_obj = format_map.get(self.backup_job.format)
            if not format_obj:
                raise ValueError(f"Unsupported format: {self.backup_job.format}")
            
            if self.backup_job.format == 'csv':
                return self._export_csv_zip(format_obj)
            else:
                return self._export_single_file(format_obj)
                
        except Exception as e:
            self.backup_job.status = 'failed'
            self.backup_job.add_log(f"Export error: {str(e)}")
            self.backup_job.save()
            raise
    
    def _export_single_file(self, format_obj):
        """Export to single file (JSON/Excel)"""
        if isinstance(format_obj, base_formats.XLSX):
            return self._export_excel()
        else:  # JSON format
            return self._export_json(format_obj)
    
    def _export_excel(self):
        """Export to Excel with multiple sheets"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for name, resource in self.resources.items():
            try:
                queryset = self._get_filtered_queryset(resource)
                if queryset and queryset.exists():
                    dataset = resource.export(queryset)
                    if dataset:
                        # Create sheet with safe name (Excel has 31 char limit)
                        sheet_name = name.replace('_', ' ').title()[:31]
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # Write headers
                        if dataset.headers:
                            for col, header in enumerate(dataset.headers, 1):
                                ws.cell(row=1, column=col, value=str(header))
                        
                        # Write data
                        for row_idx, row in enumerate(dataset.dict, 2):
                            for col_idx, (header, value) in enumerate(row.items(), 1):
                                # Handle None values and ensure proper data types
                                cell_value = value if value is not None else ''
                                if isinstance(cell_value, (list, dict)):
                                    cell_value = str(cell_value)
                                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        self.backup_job.processed_records += len(dataset)
                        self.backup_job.add_log(f"Exported {len(dataset)} {name} records")
                        
            except Exception as e:
                self.backup_job.add_log(f"Error exporting {name}: {str(e)}")
                continue
        
        # Save workbook to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()
        
        # Save file
        filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.backup_job.export_file.save(filename, ContentFile(content))
        
        self.backup_job.status = 'completed'
        self.backup_job.completed_at = datetime.now()
        self.backup_job.save()
        
        return self.backup_job.export_file
    
    def _export_json(self, format_obj):
        """Export to JSON format"""
        data = {}
        
        for name, resource in self.resources.items():
            try:
                queryset = self._get_filtered_queryset(resource)
                if queryset and queryset.exists():
                    dataset = resource.export(queryset)
                    if dataset:
                        data[name] = dataset.dict
                        self.backup_job.processed_records += len(dataset)
                        self.backup_job.add_log(f"Exported {len(dataset)} {name} records")
            except Exception as e:
                self.backup_job.add_log(f"Error exporting {name}: {str(e)}")
                continue
        
        # Convert to JSON
        import json
        content = json.dumps(data, indent=2, default=str).encode('utf-8')
        
        # Save file
        filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        self.backup_job.export_file.save(filename, ContentFile(content))
        
        self.backup_job.status = 'completed'
        self.backup_job.completed_at = datetime.now()
        self.backup_job.save()
        
        return self.backup_job.export_file
    
    def _export_csv_zip(self, format_obj):
        """Export to ZIP with multiple CSV files"""
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for name, resource in self.resources.items():
                try:
                    queryset = self._get_filtered_queryset(resource)
                    if queryset and queryset.exists():
                        dataset = resource.export(queryset)
                        if dataset:
                            csv_content = format_obj.export_data(dataset)
                            zip_file.writestr(f"{name}.csv", csv_content)
                            
                            self.backup_job.processed_records += len(dataset)
                            self.backup_job.add_log(f"Exported {len(dataset)} {name} records")
                except Exception as e:
                    self.backup_job.add_log(f"Error exporting {name}: {str(e)}")
                    continue
        
        zip_buffer.seek(0)
        filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        self.backup_job.export_file.save(filename, ContentFile(zip_buffer.getvalue()))
        
        self.backup_job.status = 'completed'
        self.backup_job.completed_at = datetime.now()
        self.backup_job.save()
        
        return self.backup_job.export_file
    
    def _get_filtered_queryset(self, resource):
        """Apply date filters to resource queryset"""
        try:
            if not hasattr(resource._meta, 'model'):
                return None
                
            queryset = resource._meta.model.objects.all()
            
            # Apply date filtering
            date_field = getattr(resource._meta, 'date_field', 'created_at')
            
            if hasattr(queryset.model, date_field):
                if self.backup_job.date_from:
                    filter_dict = {f"{date_field}__gte": self.backup_job.date_from}
                    queryset = queryset.filter(**filter_dict)
                if self.backup_job.date_to:
                    filter_dict = {f"{date_field}__lte": self.backup_job.date_to}
                    queryset = queryset.filter(**filter_dict)
            
            return queryset
        except Exception as e:
            self.backup_job.add_log(f"Error filtering queryset for {resource.__class__.__name__}: {str(e)}")
            return None
    
    def import_data(self):
        """Import data using django-import-export"""
        try:
            self.backup_job.status = 'processing'
            self.backup_job.save()
            
            format_map = {
                'json': base_formats.JSON(),
                'excel': base_formats.XLSX(),
                'csv': base_formats.CSV()
            }
            
            format_obj = format_map.get(self.backup_job.format)
            if not format_obj:
                raise ValueError(f"Unsupported format: {self.backup_job.format}")
            
            with self.backup_job.import_file.open('rb') as f:
                if self.backup_job.format == 'csv' and self.backup_job.import_file.name.endswith('.zip'):
                    return self._import_csv_zip(f)
                else:
                    return self._import_single_file(f, format_obj)
                    
        except Exception as e:
            self.backup_job.status = 'failed'
            self.backup_job.add_log(f"Import error: {str(e)}")
            self.backup_job.save()
            raise
    
    def _import_single_file(self, file_obj, format_obj):
        """Import from single file"""
        imported_count = 0
        
        if isinstance(format_obj, base_formats.XLSX):
            imported_count = self._import_excel(file_obj)
        else:  # JSON format
            imported_count = self._import_json(file_obj)
        
        self.backup_job.processed_records = imported_count
        self.backup_job.status = 'completed'
        self.backup_job.completed_at = datetime.now()
        self.backup_job.save()
        
        return imported_count
    
    def _import_excel(self, file_obj):
        """Import from Excel file"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel import")
        
        imported_count = 0
        wb = load_workbook(file_obj)
        
        for sheet_name in wb.sheetnames:
            resource_name = sheet_name.lower().replace(' ', '_')
            if resource_name in self.resources:
                resource = self.resources[resource_name]
                
                try:
                    # Convert sheet to dataset
                    ws = wb[sheet_name]
                    dataset = tablib.Dataset()
                    
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        dataset.headers = [str(header) if header else '' for header in rows[0]]
                        for row in rows[1:]:
                            # Convert None values to empty strings
                            clean_row = [val if val is not None else '' for val in row]
                            dataset.append(clean_row)
                    
                    if dataset:
                        result = resource.import_data(dataset, dry_run=False)
                        imported_count += len(result.rows)
                        self.backup_job.add_log(f"Imported {len(result.rows)} {resource_name} records")
                        
                        if result.has_errors():
                            for error in result.row_errors():
                                self.backup_job.add_log(f"Error in {resource_name}: {error}")
                                self.backup_job.error_count += 1
                                
                except Exception as e:
                    self.backup_job.add_log(f"Error importing {resource_name}: {str(e)}")
                    continue
        
        return imported_count
    
    def _import_json(self, file_obj):
        """Import from JSON file"""
        import json
        
        imported_count = 0
        content = file_obj.read().decode('utf-8')
        data = json.loads(content)
        
        for resource_name, records in data.items():
            if resource_name in self.resources:
                resource = self.resources[resource_name]
                
                try:
                    dataset = tablib.Dataset()
                    if records and isinstance(records, list):
                        # Set headers from first record
                        if records[0] and isinstance(records[0], dict):
                            dataset.headers = list(records[0].keys())
                            for record in records:
                                dataset.append([record.get(header, '') for header in dataset.headers])
                    
                    if dataset:
                        result = resource.import_data(dataset, dry_run=False)
                        imported_count += len(result.rows)
                        self.backup_job.add_log(f"Imported {len(result.rows)} {resource_name} records")
                        
                        if result.has_errors():
                            for error in result.row_errors():
                                self.backup_job.add_log(f"Error in {resource_name}: {error}")
                                self.backup_job.error_count += 1
                                
                except Exception as e:
                    self.backup_job.add_log(f"Error importing {resource_name}: {str(e)}")
                    continue
        
        return imported_count
    
    def _import_csv_zip(self, file_obj):
        """Import from ZIP with CSV files"""
        imported_count = 0
        
        with zipfile.ZipFile(file_obj, 'r') as zip_file:
            for filename in zip_file.namelist():
                if filename.endswith('.csv'):
                    resource_name = filename[:-4]  # Remove .csv extension
                    if resource_name in self.resources:
                        resource = self.resources[resource_name]
                        
                        try:
                            csv_content = zip_file.read(filename).decode('utf-8')
                            dataset = tablib.Dataset()
                            dataset.load(csv_content, format='csv')
                            
                            if dataset:
                                result = resource.import_data(dataset, dry_run=False)
                                imported_count += len(result.rows)
                                self.backup_job.add_log(f"Imported {len(result.rows)} {resource_name} records")
                                
                                if result.has_errors():
                                    for error in result.row_errors():
                                        self.backup_job.add_log(f"Error in {resource_name}: {error}")
                                        self.backup_job.error_count += 1
                                        
                        except Exception as e:
                            self.backup_job.add_log(f"Error importing {resource_name}: {str(e)}")
                            continue
        
        return imported_count